from Products.Five.browser import BrowserView
import time
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile

import openpyxl
import datetime
from openpyxl.cell import get_column_letter
from plone import api
from ejn.types import typesMessageFactory as _
import os
from openpyxl import load_workbook as open_workbook  #

import ejn.migration as base_xls_path
import transaction
from plone.namedfile.file import NamedBlobFile

from plone.dexterity.interfaces import IDexterityFTI
from zope.component import getUtility


def get_xls_file_with_result(result, headers, filename_prefix=''):
    filename = filename_prefix + 'data-%s.xlsx' % datetime.datetime.now().strftime('%Y-%m-prepared-on-%d%H%M%S')
    # self.request.response.setHeader("Content-type", "application/vnd.ms-excel")
    # self.request.response.setHeader("Content-disposition", "attachment;filename=%s" % filename)

    wbk = openpyxl.Workbook(write_only=True)
    sheet1 = wbk.create_sheet()
    sheet1.title = _('data')

    wcell = openpyxl.writer.write_only.WriteOnlyCell

    count = 0
    for row in result:
        # print row
        row_data = []
        if count == 0:
            colNo = 0
            for col in headers:
                val = col
                c1 = wcell(sheet1, value=val)
                c1.font = openpyxl.styles.Font(bold=True)
                if isinstance(val, datetime.datetime):
                    val = val.strftime('%Y-%m-%d %H:%M:%S')
                sheet1.column_dimensions[get_column_letter(colNo + 1)].width = len(val)
                row_data.append(c1)
            sheet1.append(row_data)

        row_data = []
        for col in row:
            val = col
            if isinstance(val, datetime.datetime):
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            c1 = wcell(sheet1, value=val)
            c1.font = openpyxl.styles.Font(bold=False)
            row_data.append(c1)
        sheet1.append(row_data)

        count += 1

    wbk.save(filename='/tmp/%s' % filename)
    fp = open('/tmp/%s' % filename)
    data = fp.read()
    return data


class DownloadApplications(BrowserView):
    render = ViewPageTemplateFile('download_applications.pt')

    def __call__(self):

        if self.context.REQUEST.get('download', '') == 'yes':
            filename = self.context.getId() + '.xlsx'
            self.request.response.setHeader("Content-type", "application/vnd.ms-excel")
            self.request.response.setHeader("Content-disposition", "attachment;filename=%s" % filename)
            return self.download_applications()
        return self.render()

    def stringify_value(self, val, field, app):
        if hasattr(val, 'raw'):
            return val.raw
        if isinstance(val, NamedBlobFile):
            return app.absolute_url() + '/@@download/' + field
        if isinstance(val, set):
            return ', '.join(val)
        if isinstance(val, list):
            return ', '.join(val)
        return val

    def download_applications(self):
        """
        Generate a CSV file with all the users registered before 2013 by 
        scrapping their email addresses based on crossing user credentials and the associated user profiles with URLs that has only numbers after the last slash;
        """
        result = []
        headers = []
        applications = self.context.objectValues()
        count = 0
        total = len(applications)
        limit = 100
        for app in applications:
            count += 1
            if len(headers) == 0:
                schema = getUtility(IDexterityFTI, name=app.portal_type).lookupSchema()
                headers = schema.names()

            # import pdb;pdb.set_trace()
            row = []
            for field in headers:
                val = getattr(app, field)
                val = self.stringify_value(val, field, app)
                row.append(val)
            result.append(row)
        data = get_xls_file_with_result(result=result, headers=headers)
        return data
