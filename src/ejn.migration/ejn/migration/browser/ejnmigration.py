from Products.Five.browser import BrowserView
import time
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile

import openpyxl
import datetime
from openpyxl.cell import get_column_letter
from plone import api
from ejn.migration import migrationMessageFactory as _
import os
from openpyxl import load_workbook as open_workbook  #

import ejn.migration as base_xls_path
import transaction
from ejn.types.content.memberprofile import MemberProfileSchema

def get_xls_file_with_result_X(result, headers, filename_prefix=''):


    import csv
    filename = '/tmp/' + filename_prefix + 'data-%s.csv' % datetime.datetime.now().strftime('%Y-%m-prepared-on-%d%H%M%S')

    with open(filename, mode='w') as employee_file:
        employee_writer = csv.writer(employee_file, delimiter='$', quotechar='"', quoting=csv.QUOTE_ALL)
        employee_writer.writerow(headers)
        for row in result:
            # import pdb;pdb.set_trace()
            employee_writer.writerow(row)

    fp = open('%s' % filename)
    data = fp.read()
    return data


def get_xls_file_with_result(result, headers, filename_prefix=''):
    filename = filename_prefix + 'data-%s.xlsx' % datetime.datetime.now().strftime('%Y-%m-prepared-on-%d%H%M%S')
    # self.request.response.setHeader("Content-type", "application/vnd.ms-excel")
    # self.request.response.setHeader("Content-disposition", "attachment;filename=%s" % filename)

    wbk = openpyxl.Workbook(write_only=True)
    sheet1 = wbk.create_sheet()
    sheet1.title = _('data')

    wcell = openpyxl.writer.write_only.WriteOnlyCell

    count = 0
    for row in result:
        # print row
        row_data = []
        if count == 0:
            colNo = 0
            for col in headers:
                val = col
                c1 = wcell(sheet1, value=val)
                c1.font = openpyxl.styles.Font(bold=True)
                if isinstance(val, datetime.datetime):
                    val = val.strftime('%Y-%m-%d %H:%M:%S')
                sheet1.column_dimensions[get_column_letter(colNo + 1)].width = len(val)
                row_data.append(c1)
            sheet1.append(row_data)

        row_data = []
        for col in row:
            val = col
            if isinstance(val, datetime.datetime):
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            c1 = wcell(sheet1, value=val)
            c1.font = openpyxl.styles.Font(bold=False)
            row_data.append(c1)
        sheet1.append(row_data)

        count += 1

    wbk.save(filename='/tmp/%s' % filename)
    fp = open('/tmp/%s' % filename)
    data = fp.read()
    return data

import six

def make_smart_text(text, encoding='utf-8', errors='strict'):
    try:
        return text.decode('iso-8859-1').encode(encoding)
    except:
        try:
           return text.decode('utf-8').encode(encoding)
        except: 
            return text
    return text


class EjnMigration(BrowserView):
    render = ViewPageTemplateFile('ejn_migration.pt')

    def __call__(self):
        if self.context.REQUEST.get('download_users', '')  == 'yes':
            filename = 'all_users_with_profile.xlsx'
            self.request.response.setHeader("Content-type", "application/vnd.ms-excel")
            self.request.response.setHeader("Content-disposition", "attachment;filename=%s" % filename)
            return self.run_download_users()
        if self.context.REQUEST.get('run_step', '') == 'step1':
            filename = 'step1_file.xlsx'
            self.request.response.setHeader("Content-type", "application/vnd.ms-excel")
            self.request.response.setHeader("Content-disposition", "attachment;filename=%s" % filename)
            return self.run_step_one()
        if self.context.REQUEST.get('delete_users_marked_yes_in_xls', '') == 'yes':
            self.delete_users_marked_yes_in_xls()
            return 'Done'

        if self.context.REQUEST.get('download_content', '')  == 'yes':
            filename = 'all_content.csv'
            self.request.response.setHeader("Content-type", "text/csv")
            self.request.response.setHeader("Content-disposition", "attachment;filename=%s" % filename)
            return self.run_download_all_content()
        return self.render()

    def run_download_all_content(self):
        if self.context.REQUEST.get('type', '')in ['Story', 'Program Update', 'Reporter Resource', 'Program', 'Opportunity', 'Document']:
            result = api.content.find(context=self.context, portal_type=self.context.REQUEST.get('type', ''))
            xldata = []
            # import pdb;pdb.set_trace()
            headers = []
            fields = []
            for row in result:
                data_row = []
                obj = row.getObject()
                data_row.append(obj.absolute_url().replace('dev.earthjournalism.net', 'earthjournalism.net'))
                if len(fields) == 0:
                    fields = obj.schema.fields()
                    headers = ['URL'] + [x.widget.label for x in fields]
                for field in fields:
                    val = getattr(obj, field.accessor)()
                    data_row.append(str(val))
                xldata.append(data_row)
        data = get_xls_file_with_result_X(result=xldata, headers=headers)
        return data


    def run_download_users(self):
        result = []
        headers = ['fullname', 'email', 'gender', 'country']
        # member_fields = MemberProfileSchema.fields()
        # for field in member_fields:
        #    headers.append(field.getName())

        users = api.user.get_users()
        count = 0
        total = len(users)
        limit = 100
        for user in users:
            count += 1
            fname = user.getProperty('fullname')
            if fname is None:
                fname = user.getId()
            fname = make_smart_text(fname)
            result_row = [fname, user.getProperty('email'), user.getProperty('gender'), user.getProperty('country')]
            result.append(result_row)
            print count, total
            if 1 == 2:
                # import pdb;pdb.set_trace()
                # if count > limit:
                #    break
                user_profile = self.get_user_profile(user, return_obj=True)

                print count, total, user_profile
                if user_profile:
                    user_profile_link = user_profile.absolute_url()
                else:
                    user_profile_link = ''
                url = user.absolute_url().replace('http://localhost:8080/Plone/', 'https://www.earthjournalism.net/')
                user_profile_link = user_profile_link.replace('http://localhost:8080/Plone/', 'https://www.earthjournalism.net/')
                last_login_time = user.getProperty('last_login_time')
                if last_login_time:
                    if hasattr(last_login_time, 'strftime'):
                        last_login_time = last_login_time.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        last_login_time = str(last_login_time)
                else:
                    last_login_time = str(last_login_time)
                result_row = [user.getId(), user.getProperty('email'), url, user_profile_link, last_login_time]
                for field in member_fields:
                    if user_profile:
                        val = str(getattr(user_profile, field.accessor)())
                    else:
                        val = ''
                    result_row.append(val)
                result.append(result_row)
        data = get_xls_file_with_result(result=result, headers=headers)
        return data

    def get_user_profile(self, member, return_obj=False):
        # from Products.CMFCore.utils import getToolByName

        cat = self.context.portal_catalog

        # member = context.portal_membership.getAuthenticatedMember()

        userid = member.getUserName()
        # userid = member.id
        email = member.getProperty('email')

        results_by_userid = cat.searchResults(portal_type='Member Profile', Creator=userid)
        results_by_email = cat.searchResults(portal_type='Member Profile', email=email)

        matches_by_userid = len(results_by_userid)
        matches_by_email = len(results_by_email)

        if matches_by_userid >= 1:
            profile = results_by_userid[0]
            if return_obj:
                return profile.getObject()
            reurl = profile.getObject().absolute_url()
        elif matches_by_email >= 1:
            profile = results_by_email[0]
            if return_obj:
                return profile.getObject()
            reurl = profile.getObject().absolute_url()
        else:
            reurl = ''
        if return_obj:
            return None
        return reurl

    def run_step_one(self):
        """
        Generate a CSV file with all the users registered before 2013 by 
        scrapping their email addresses based on crossing user credentials and the associated user profiles with URLs that has only numbers after the last slash;
        """
        result = []
        headers = ['username', 'email', 'user_link', 'user_profile_link', 'last_login_time']
        users = api.user.get_users()
        count = 0
        total = len(users)
        limit = 100
        for user in users:
            count += 1
            # import pdb;pdb.set_trace()
            # if count > limit:
            #    break
            user_profile_link = self.get_user_profile(user)
            print count, total, user_profile_link
            url = user.absolute_url().replace('http://localhost:8080/Plone/', 'https://www.earthjournalism.net/')
            user_profile_link = user_profile_link.replace('http://localhost:8080/Plone/', 'https://www.earthjournalism.net/')
            last_login_time = user.getProperty('last_login_time')
            if last_login_time:
                if hasattr(last_login_time, 'strftime'):
                    last_login_time = last_login_time.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    last_login_time = str(last_login_time)
            else:
                last_login_time = str(last_login_time)
            result.append([user.getId(), user.getProperty('email'), url, user_profile_link, last_login_time])
        data = get_xls_file_with_result(result=result, headers=headers)
        return data

    def remove_user(self, email):
        """Remove subscriber. We can not use `api.user.delete` as, by
        default, it tries to remove member areas and local roles; the
        later consumes a lot of memory and is not necessary for us.
        """
        membership_tool = api.portal.get_tool('portal_membership')
        with api.env.adopt_roles(['Manager']):
            membership_tool.deleteMembers(
                [email], delete_memberareas=True, delete_localroles=False)

    def delete_users_marked_yes_in_xls(self):
        dir_name_old_vdex = os.path.dirname(base_xls_path.__file__)
        pjoin = os.path.join
        old_vdex_path = pjoin(dir_name_old_vdex, 'browser')
        book = open_workbook(old_vdex_path + '/step1_file_with_last_login_time_dev2_reviewed.xlsx')
        sheet = book.get_sheet_by_name('data')
        usernames_to_delete = {}
        for index, row in enumerate(sheet.rows):
            if index == 0:
                continue
            username = row[0].value
            # import pdb;pdb.set_trace()
            if not usernames_to_delete.has_key(username):
                if row[5].value == 'yes':
                    usernames_to_delete[username] = [row[0].value, row[2].value, row[3].value, row[5].value]
        # import pdb;pdb.set_trace()
        self.context.plone_log('geting users...')
        users = api.user.get_users()
        usernames_to_delete_names = usernames_to_delete.keys()
        count = 1
        total = len(users)
        self.context.plone_log(total)
        portal = api.portal.get()
        time_start = time.time()
        for userobj in users:
            count += 1
            if userobj.getId() in usernames_to_delete_names:

                # import pdb;pdb.set_trace()
                # if count < 62000:
                #    self.context.plone_log('skip %s' % str(count))
                #    continue
                self.remove_user(email=userobj.getId())
                profile_link = usernames_to_delete.get(userobj.getId())[2]
                if profile_link:
                    profile_paths = profile_link.split('https://www.earthjournalism.net/directory/')
                    profile_path = profile_paths[1]
                    obj_profile = portal['directory'][profile_path]
                    api.content.delete(obj=obj_profile)
                    self.context.plone_log([userobj, obj_profile])
                if count % 100 == 0:
                    self.context.plone_log('done %s out of total %s' % (str(count), str(total)))
                    deltatime = time.time() - time_start
                    deltatime_minutes = deltatime / 60
                    percentage = float(count) / total
                    compleetion_time_minutes = 1 / percentage * deltatime / 60 - deltatime_minutes
                    self.context.plone_log("timedelta minutes: %s  number %s of %s expected compleetion in %s minutes" %
                                           (deltatime_minutes, count, total, compleetion_time_minutes))
                    transaction.commit()
        return usernames_to_delete
